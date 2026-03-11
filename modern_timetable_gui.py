import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import sys
import io
import threading
import sqlite3
import os

# helper that respects environment variable in db_operations

def _connect():
    from db_operations import DB_PATH
    return sqlite3.connect(DB_PATH)
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from db_operations import initialize_database, fetch_data_from_db, get_course_classes, set_course_classes, DB_PATH
from db_operations import add_professor, add_course, add_branch, add_branch_teacher_course_mapping
from db_operations import delete_course, delete_course_teacher_mapping
from db_operations import get_all_professors, get_professor_courses, get_available_courses_for_professor
from db_operations import add_course_to_professor, remove_course_from_professor
from simplified_genetic_algorithm import genetic_algorithm, print_timetable, print_professor_schedules
from main import debug_database, verify_database_integrity

# Color scheme
COLORS = {
    "primary": "#3498db",      # Blue
    "secondary": "#2ecc71",    # Green
    "accent": "#e74c3c",       # Red
    "background": "#f5f5f5",   # Light gray
    "text": "#2c3e50",         # Dark blue/gray
    "light_text": "#7f8c8d",   # Light gray text
    "success": "#27ae60",      # Dark green
    "warning": "#f39c12",      # Orange
    "error": "#c0392b"         # Dark red
}

# Redirect stdout to capture print statements
class StdoutRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget
        self.buffer = io.StringIO()

    def write(self, string):
        self.buffer.write(string)
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)  # Auto-scroll to the end
        
    def flush(self):
        pass

class ModernTimetableApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Fast Timetable Generator")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 700)
        
        # Configure the style
        self.configure_style()
        
        # Store the last generated timetable and professor schedules
        self.last_timetable = None
        self.last_professor_schedules = None
        
        # Create main frame
        self.main_frame = ttk.Frame(root)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create sidebar and content area
        self.create_sidebar()
        self.create_content_area()
        
        # Initialize the database
        self.initialize_db()
        
        # Show the home screen by default
        self.show_home()
    
    def configure_style(self):
        """Configure the ttk styles for a modern look"""
        style = ttk.Style()
        style.theme_use('clam')  # Use the 'clam' theme as a base
        
        # Configure colors
        style.configure('TFrame', background=COLORS["background"])
        style.configure('TLabel', background=COLORS["background"], foreground=COLORS["text"])
        style.configure('TButton', background=COLORS["primary"], foreground="white", font=('Arial', 10))
        style.map('TButton', background=[('active', COLORS["secondary"])])
        
        # Sidebar button style
        style.configure('Sidebar.TButton', font=('Arial', 11), padding=10)
        
        # Heading style
        style.configure('Heading.TLabel', font=('Arial', 16, 'bold'), foreground=COLORS["primary"])
        style.configure('SubHeading.TLabel', font=('Arial', 12, 'italic'), foreground=COLORS["light_text"])
        
        # Card style
        style.configure('Card.TFrame', background="white", relief="raised", borderwidth=1)
        
        # Success and error styles
        style.configure('Success.TLabel', foreground=COLORS["success"])
        style.configure('Error.TLabel', foreground=COLORS["error"])
    
    def create_sidebar(self):
        """Create the sidebar with navigation buttons"""
        self.sidebar = ttk.Frame(self.main_frame, width=200, style='TFrame')
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y, padx=0, pady=0)
        
        # Make the sidebar fixed width
        self.sidebar.pack_propagate(False)
        
        # App title
        title_frame = ttk.Frame(self.sidebar, style='TFrame')
        title_frame.pack(fill=tk.X, padx=10, pady=20)
        
        ttk.Label(title_frame, text="Timetable", font=('Arial', 18, 'bold'), 
                 foreground=COLORS["primary"], style='TLabel').pack(anchor=tk.W)
        ttk.Label(title_frame, text="Generator", font=('Arial', 14), 
                 foreground=COLORS["secondary"], style='TLabel').pack(anchor=tk.W)
        
        # Navigation buttons
        self.nav_buttons = []
        
        nav_items = [
            ("Home", self.show_home),
            ("Database", self.show_database),
            ("Courses", self.show_courses),
            ("Classes", self.show_classes),
            ("Professor Courses", self.show_professor_courses),
            ("Generate", self.show_generate),
            ("View Timetable", self.show_timetable),
            ("Professor Schedules", self.show_professors)
        ]
        
        for text, command in nav_items:
            btn = ttk.Button(self.sidebar, text=text, command=command, style='Sidebar.TButton')
            btn.pack(fill=tk.X, padx=10, pady=5)
            self.nav_buttons.append(btn)
        
        # Status section at bottom of sidebar
        status_frame = ttk.Frame(self.sidebar, style='TFrame')
        status_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=20)
        
        ttk.Label(status_frame, text="Status:", style='TLabel').pack(anchor=tk.W)
        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(status_frame, textvariable=self.status_var, style='TLabel').pack(anchor=tk.W, pady=5)
    
    def create_content_area(self):
        """Create the main content area"""
        self.content = ttk.Frame(self.main_frame, style='TFrame')
        self.content.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Create frames for each section (only one will be visible at a time)
        self.home_frame = ttk.Frame(self.content, style='TFrame')
        self.database_frame = ttk.Frame(self.content, style='TFrame')
        self.courses_frame = ttk.Frame(self.content, style='TFrame')
        self.classes_frame = ttk.Frame(self.content, style='TFrame')  # New frame for classes management
        self.professor_courses_frame = ttk.Frame(self.content, style='TFrame')  # New frame for professor courses
        self.generate_frame = ttk.Frame(self.content, style='TFrame')
        self.timetable_frame = ttk.Frame(self.content, style='TFrame')
        self.professors_frame = ttk.Frame(self.content, style='TFrame')

        # Set up each frame
        self.setup_home_frame()
        self.setup_database_frame()
        self.setup_courses_frame()
        self.setup_classes_frame()  # Setup the new classes frame
        self.setup_professor_courses_frame()  # Setup the professor courses frame
        self.setup_generate_frame()
        self.setup_timetable_frame()
        self.setup_professors_frame()
    
    def hide_all_frames(self):
        """Hide all content frames"""
        for frame in [self.home_frame, self.database_frame, self.courses_frame, self.classes_frame,
                     self.professor_courses_frame, self.generate_frame, self.timetable_frame, self.professors_frame]:
            frame.pack_forget()
    
    def show_home(self):
        self.hide_all_frames()
        self.home_frame.pack(fill=tk.BOTH, expand=True)
    
    def show_database(self):
        self.hide_all_frames()
        self.database_frame.pack(fill=tk.BOTH, expand=True)
    
    def show_courses(self):
        self.hide_all_frames()
        self.courses_frame.pack(fill=tk.BOTH, expand=True)
        self.refresh_courses()
        self.refresh_branches()
        self.refresh_mapping_lists()

    def show_classes(self):
        self.hide_all_frames()
        self.classes_frame.pack(fill=tk.BOTH, expand=True)
        self.refresh_courses_for_classes()

    def show_professor_courses(self):
        self.hide_all_frames()
        self.professor_courses_frame.pack(fill=tk.BOTH, expand=True)
        self.refresh_professors_list()

    def show_generate(self):
        self.hide_all_frames()
        self.generate_frame.pack(fill=tk.BOTH, expand=True)
    
    def show_timetable(self):
        self.hide_all_frames()
        self.timetable_frame.pack(fill=tk.BOTH, expand=True)
    
    def show_professors(self):
        self.hide_all_frames()
        self.professors_frame.pack(fill=tk.BOTH, expand=True)
        self.refresh_professors()
    
    def setup_home_frame(self):
        """Set up the home screen with a modern, clean design"""
        # Title with improved styling
        title_frame = ttk.Frame(self.home_frame, style='TFrame')
        title_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Label(title_frame, text="Timetable Generator",
                 font=('Arial', 24, 'bold'), foreground=COLORS["primary"],
                 style='TLabel').pack(pady=(0, 5))
        ttk.Label(title_frame, text="Create optimal timetables using genetic algorithms",
                 style='SubHeading.TLabel').pack(pady=(0, 10))

        # Main content in a two-column layout
        content_frame = ttk.Frame(self.home_frame, style='TFrame')
        content_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        content_frame.columnconfigure(0, weight=1)
        content_frame.columnconfigure(1, weight=1)

        # Left column: Stats and actions
        left_col = ttk.Frame(content_frame, style='TFrame')
        left_col.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N+tk.S+tk.W+tk.E)

        # Stats cards
        stats_frame = ttk.LabelFrame(left_col, text="Database Statistics", style='TFrame')
        stats_frame.pack(fill=tk.X, pady=10)

        stats_grid = ttk.Frame(stats_frame, style='TFrame')
        stats_grid.pack(fill=tk.X, padx=10, pady=10)
        stats_grid.columnconfigure(0, weight=1)
        stats_grid.columnconfigure(1, weight=1)
        stats_grid.columnconfigure(2, weight=1)

        # Create stat cards in a grid
        self.create_stat_card(stats_grid, "Branches", "branch_count", 0, 0)
        self.create_stat_card(stats_grid, "Courses", "course_count", 0, 1)
        self.create_stat_card(stats_grid, "Teachers", "teacher_count", 0, 2)

        # Quick actions
        actions_frame = ttk.LabelFrame(left_col, text="Quick Actions", style='TFrame')
        actions_frame.pack(fill=tk.X, pady=10)

        # Create a grid for buttons with 2 columns
        button_grid = ttk.Frame(actions_frame, style='TFrame')
        button_grid.pack(fill=tk.X, padx=10, pady=10)
        button_grid.columnconfigure(0, weight=1)
        button_grid.columnconfigure(1, weight=1)

        # Add buttons with icons (using text for now)
        ttk.Button(button_grid, text="Initialize Database",
                  command=self.initialize_db).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W+tk.E)
        ttk.Button(button_grid, text="Generate Timetable",
                  command=self.generate_timetable).grid(row=0, column=1, padx=5, pady=5, sticky=tk.W+tk.E)
        ttk.Button(button_grid, text="View Timetable",
                  command=self.show_timetable).grid(row=1, column=0, padx=5, pady=5, sticky=tk.W+tk.E)
        ttk.Button(button_grid, text="View Professor Schedules",
                  command=self.show_professors).grid(row=1, column=1, padx=5, pady=5, sticky=tk.W+tk.E)

        # Right column: Console output
        right_col = ttk.Frame(content_frame, style='TFrame')
        right_col.grid(row=0, column=1, padx=10, pady=10, sticky=tk.N+tk.S+tk.W+tk.E)

        console_frame = ttk.LabelFrame(right_col, text="Console Output", style='TFrame')
        console_frame.pack(fill=tk.BOTH, expand=True)

        self.console = scrolledtext.ScrolledText(console_frame, wrap=tk.WORD, height=15,
                                               font=('Consolas', 9))
        self.console.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Add a welcome message to the console
        welcome_msg = (
            "Welcome to Timetable Generator!\n\n"
            "This application helps you create optimal timetables using genetic algorithms.\n\n"
            "To get started:\n"
            "1. Initialize the database\n"
            "2. Add branches, professors, and courses\n"
            "3. Generate a timetable\n\n"
            "System ready and waiting for commands...\n"
        )
        self.console.insert(tk.END, welcome_msg)

        # Redirect stdout to the console
        sys.stdout = StdoutRedirector(self.console)
    
    def create_stat_card(self, parent, title, var_name, row, col):
        """Create a statistics card"""
        card = ttk.Frame(parent, style='Card.TFrame')
        card.grid(row=row, column=col, padx=10, pady=10, sticky=tk.W+tk.E)
        
        ttk.Label(card, text=title, font=('Arial', 12), background="white").pack(pady=(10, 5))
        
        # Create a variable to hold the count
        setattr(self, var_name, tk.StringVar(value="0"))
        ttk.Label(card, textvariable=getattr(self, var_name), 
                 font=('Arial', 24, 'bold'), foreground=COLORS["primary"], 
                 background="white").pack(pady=(0, 10))
    
    def setup_database_frame(self):
        """Set up the database management frame"""
        # Title
        ttk.Label(self.database_frame, text="Database Management", 
                 style='Heading.TLabel').pack(pady=(0, 20))
        
        # Create a notebook for tabs
        notebook = ttk.Notebook(self.database_frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # Create tabs
        actions_tab = ttk.Frame(notebook, style='TFrame')
        branch_tab = ttk.Frame(notebook, style='TFrame')
        professor_tab = ttk.Frame(notebook, style='TFrame')
        
        notebook.add(actions_tab, text="Database Actions")
        notebook.add(branch_tab, text="Add Branch")
        notebook.add(professor_tab, text="Add Professor")
        
        # Database actions tab
        actions_frame = ttk.Frame(actions_tab, style='TFrame')
        actions_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        ttk.Button(actions_frame, text="Initialize Database", 
                  command=self.initialize_db).pack(fill=tk.X, pady=10)
        ttk.Button(actions_frame, text="Debug Database", 
                  command=self.debug_database).pack(fill=tk.X, pady=10)
        ttk.Button(actions_frame, text="Verify Database Integrity", 
                  command=self.verify_database).pack(fill=tk.X, pady=10)
        
        # Console output for actions
        console_frame = ttk.LabelFrame(actions_frame, text="Console Output", style='TFrame')
        console_frame.pack(fill=tk.BOTH, expand=True, pady=20)
        
        self.db_console = scrolledtext.ScrolledText(console_frame, wrap=tk.WORD, height=10)
        self.db_console.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Branch tab
        branch_frame = ttk.Frame(branch_tab, style='TFrame')
        branch_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        ttk.Label(branch_frame, text="Branch Name:").pack(anchor=tk.W, pady=(0, 5))
        self.branch_name_var = tk.StringVar()
        ttk.Entry(branch_frame, textvariable=self.branch_name_var, width=40).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(branch_frame, text="Add Branch", 
                  command=self.add_branch).pack(anchor=tk.W, pady=10)
        
        # Professor tab
        prof_frame = ttk.Frame(professor_tab, style='TFrame')
        prof_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        ttk.Label(prof_frame, text="Professor Name:").pack(anchor=tk.W, pady=(0, 5))
        self.prof_name_var = tk.StringVar()
        ttk.Entry(prof_frame, textvariable=self.prof_name_var, width=40).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(prof_frame, text="Branch:").pack(anchor=tk.W, pady=(10, 5))
        self.prof_branch_var = tk.StringVar()
        self.prof_branch_combo = ttk.Combobox(prof_frame, textvariable=self.prof_branch_var, width=40)
        self.prof_branch_combo.pack(fill=tk.X, pady=(0, 10))
        
        button_frame = ttk.Frame(prof_frame, style='TFrame')
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="Add Professor", 
                  command=self.add_professor).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Refresh Branches", 
                  command=self.refresh_branches).pack(side=tk.LEFT)
    
    def setup_courses_frame(self):
        """Set up the course management frame"""
        # Title
        ttk.Label(self.courses_frame, text="Course Management", 
                 style='Heading.TLabel').pack(pady=(0, 20))
        
        # Create a notebook for tabs
        notebook = ttk.Notebook(self.courses_frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # Create tabs
        add_course_tab = ttk.Frame(notebook, style='TFrame')
        mapping_tab = ttk.Frame(notebook, style='TFrame')
        delete_tab = ttk.Frame(notebook, style='TFrame')
        
        notebook.add(add_course_tab, text="Add Course")
        notebook.add(mapping_tab, text="Add Mapping")
        notebook.add(delete_tab, text="Delete Course")
        
        # Add course tab
        course_frame = ttk.Frame(add_course_tab, style='TFrame')
        course_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        ttk.Label(course_frame, text="Course Code:").pack(anchor=tk.W, pady=(0, 5))
        self.course_code_var = tk.StringVar()
        ttk.Entry(course_frame, textvariable=self.course_code_var, width=40).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(course_frame, text="Course Name:").pack(anchor=tk.W, pady=(10, 5))
        self.course_name_var = tk.StringVar()
        ttk.Entry(course_frame, textvariable=self.course_name_var, width=40).pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(course_frame, text="Branch:").pack(anchor=tk.W, pady=(10, 5))
        self.course_branch_var = tk.StringVar()
        self.course_branch_combo = ttk.Combobox(course_frame, textvariable=self.course_branch_var, width=40)
        self.course_branch_combo.pack(fill=tk.X, pady=(0, 10))
        
        button_frame = ttk.Frame(course_frame, style='TFrame')
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="Add Course", 
                  command=self.add_course).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Refresh Branches", 
                  command=self.refresh_branches).pack(side=tk.LEFT)
        
        # Mapping tab
        mapping_frame = ttk.Frame(mapping_tab, style='TFrame')
        mapping_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        ttk.Label(mapping_frame, text="Branch:").pack(anchor=tk.W, pady=(0, 5))
        self.mapping_branch_var = tk.StringVar()
        self.mapping_branch_combo = ttk.Combobox(mapping_frame, textvariable=self.mapping_branch_var, width=40)
        self.mapping_branch_combo.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(mapping_frame, text="Teacher:").pack(anchor=tk.W, pady=(10, 5))
        self.mapping_teacher_var = tk.StringVar()
        self.mapping_teacher_combo = ttk.Combobox(mapping_frame, textvariable=self.mapping_teacher_var, width=40)
        self.mapping_teacher_combo.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(mapping_frame, text="Course:").pack(anchor=tk.W, pady=(10, 5))
        self.mapping_course_var = tk.StringVar()
        self.mapping_course_combo = ttk.Combobox(mapping_frame, textvariable=self.mapping_course_var, width=40)
        self.mapping_course_combo.pack(fill=tk.X, pady=(0, 10))
        
        button_frame = ttk.Frame(mapping_frame, style='TFrame')
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="Add Mapping", 
                  command=self.add_mapping).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Refresh Lists", 
                  command=self.refresh_mapping_lists).pack(side=tk.LEFT)
        
        # Delete course tab
        delete_frame = ttk.Frame(delete_tab, style='TFrame')
        delete_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        ttk.Label(delete_frame, text="Course:").pack(anchor=tk.W, pady=(0, 5))
        self.delete_course_var = tk.StringVar()
        self.delete_course_combo = ttk.Combobox(delete_frame, textvariable=self.delete_course_var, width=40)
        self.delete_course_combo.pack(fill=tk.X, pady=(0, 10))
        
        button_frame = ttk.Frame(delete_frame, style='TFrame')
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="Delete Course", 
                  command=self.delete_course).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Refresh Courses", 
                  command=self.refresh_courses).pack(side=tk.LEFT)
        
        # Console output for all tabs
        console_frame = ttk.LabelFrame(self.courses_frame, text="Console Output", style='TFrame')
        console_frame.pack(fill=tk.BOTH, expand=True, pady=20)
        
        self.course_console = scrolledtext.ScrolledText(console_frame, wrap=tk.WORD, height=10)
        self.course_console.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
    
    def setup_generate_frame(self):
        """Set up the timetable generation frame"""
        # Title
        ttk.Label(self.generate_frame, text="Generate Timetable",
                 style='Heading.TLabel').pack(pady=(0, 20))
        ttk.Label(self.generate_frame, text="Using genetic algorithm for optimal timetable generation",
                 style='SubHeading.TLabel').pack(pady=(0, 10))

        # Generation options
        options_frame = ttk.LabelFrame(self.generate_frame, text="Generation Options", style='TFrame')
        options_frame.pack(fill=tk.X, pady=10)

        # Create a grid layout with 2 columns
        options_frame.columnconfigure(0, weight=0)
        options_frame.columnconfigure(1, weight=1)

        # Generations option
        ttk.Label(options_frame, text="Generations:").grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)
        self.generations_var = tk.StringVar(value="100")
        ttk.Spinbox(options_frame, from_=30, to=300, increment=10,
                   textvariable=self.generations_var, width=10).grid(row=0, column=1, padx=10, pady=10, sticky=tk.W)

        # Generate button
        generate_btn = ttk.Button(options_frame, text="Generate Timetable",
                                 command=self.generate_timetable)
        generate_btn.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky=tk.W+tk.E)

        # Status indicator
        status_frame = ttk.Frame(self.generate_frame, style='TFrame')
        status_frame.pack(fill=tk.X, pady=10)

        self.progress_var = tk.StringVar(value="Not started")
        ttk.Label(status_frame, text="Status:").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Label(status_frame, textvariable=self.progress_var,
                 font=('Arial', 10, 'bold')).pack(side=tk.LEFT)

        # Information panel
        info_frame = ttk.LabelFrame(self.generate_frame, text="Information", style='TFrame')
        info_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        info_text = (
            "The timetable generation process will run in the background.\n\n"
            "Once complete, you will be automatically redirected to the timetable view.\n\n"
            "For large datasets, this process may take several minutes.\n"
            "Please be patient while the genetic algorithm optimizes your timetable."
        )

        ttk.Label(info_frame, text=info_text, wraplength=500, justify="center").pack(
            fill=tk.BOTH, expand=True, padx=20, pady=20
        )
    
    def setup_timetable_frame(self):
        """Set up the timetable display frame"""
        # Title
        ttk.Label(self.timetable_frame, text="Timetable View",
                 style='Heading.TLabel').pack(pady=(0, 20))

        # Actions toolbar
        actions_frame = ttk.Frame(self.timetable_frame, style='TFrame')
        actions_frame.pack(fill=tk.X, pady=10)

        # Create a more organized button layout
        button_frame = ttk.Frame(actions_frame, style='TFrame')
        button_frame.pack(fill=tk.X)

        # Configure columns for even spacing
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)
        button_frame.columnconfigure(2, weight=1)

        # Add buttons in a grid layout
        ttk.Button(button_frame, text="Generate New Timetable",
                  command=self.show_generate).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W+tk.E)
        ttk.Button(button_frame, text="View Professor Schedules",
                  command=self.show_professors).grid(row=0, column=1, padx=5, pady=5, sticky=tk.W+tk.E)
        ttk.Button(button_frame, text="Export to Excel",
                  command=self.export_timetable_to_excel).grid(row=0, column=2, padx=5, pady=5, sticky=tk.W+tk.E)

        # Timetable display with improved styling
        timetable_frame = ttk.LabelFrame(self.timetable_frame, text="Current Timetable", style='TFrame')
        timetable_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        # Use a monospaced font for better timetable display
        self.timetable_text = scrolledtext.ScrolledText(
            timetable_frame,
            wrap=tk.WORD,
            font=('Courier New', 10)
        )
        self.timetable_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Initial message with better formatting
        self.timetable_text.insert(tk.END,
            "No timetable has been generated yet.\n\n"
            "Please go to the Generate tab to create a new timetable."
        )
    
    def setup_classes_frame(self):
        """Set up the course classes management frame"""
        # Title
        ttk.Label(self.classes_frame, text="Course Classes Management",
                 style='Heading.TLabel').pack(pady=(0, 20))
        ttk.Label(self.classes_frame, text="Set the number of classes per week for each course",
                 style='SubHeading.TLabel').pack(pady=(0, 20))

        # Create a notebook for tabs
        notebook = ttk.Notebook(self.classes_frame)
        notebook.pack(fill=tk.BOTH, expand=True)

        # Create tabs
        view_tab = ttk.Frame(notebook, style='TFrame')
        update_tab = ttk.Frame(notebook, style='TFrame')
        update_all_tab = ttk.Frame(notebook, style='TFrame')

        notebook.add(view_tab, text="View Classes")
        notebook.add(update_tab, text="Update Course")
        notebook.add(update_all_tab, text="Update All Courses")

        # View classes tab
        view_frame = ttk.Frame(view_tab, style='TFrame')
        view_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        ttk.Button(view_frame, text="View Course Classes Configuration",
                  command=self.view_course_classes).pack(fill=tk.X, pady=10)

        # Update single course tab
        update_frame = ttk.Frame(update_tab, style='TFrame')
        update_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        ttk.Label(update_frame, text="Course:").pack(anchor=tk.W, pady=(0, 5))
        self.update_course_var = tk.StringVar()
        self.update_course_combo = ttk.Combobox(update_frame, textvariable=self.update_course_var, width=40)
        self.update_course_combo.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(update_frame, text="Classes per Week:").pack(anchor=tk.W, pady=(10, 5))
        self.update_classes_var = tk.StringVar(value="2")
        ttk.Spinbox(update_frame, from_=1, to=5, textvariable=self.update_classes_var, width=10).pack(anchor=tk.W, pady=(0, 10))

        button_frame = ttk.Frame(update_frame, style='TFrame')
        button_frame.pack(fill=tk.X, pady=10)

        ttk.Button(button_frame, text="Update Course",
                  command=self.update_course_classes).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Refresh Courses",
                  command=self.refresh_courses_for_classes).pack(side=tk.LEFT)

        # Update all courses tab
        update_all_frame = ttk.Frame(update_all_tab, style='TFrame')
        update_all_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        ttk.Label(update_all_frame, text="Classes per Week for ALL courses:").pack(anchor=tk.W, pady=(0, 5))
        self.update_all_classes_var = tk.StringVar(value="2")
        ttk.Spinbox(update_all_frame, from_=1, to=5, textvariable=self.update_all_classes_var, width=10).pack(anchor=tk.W, pady=(0, 10))

        ttk.Button(update_all_frame, text="Update All Courses",
                  command=self.update_all_course_classes).pack(anchor=tk.W, pady=10)

        # Console output for all tabs
        console_frame = ttk.LabelFrame(self.classes_frame, text="Console Output", style='TFrame')
        console_frame.pack(fill=tk.BOTH, expand=True, pady=20)

        self.classes_console = scrolledtext.ScrolledText(console_frame, wrap=tk.WORD, height=10)
        self.classes_console.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def setup_professor_courses_frame(self):
        """Set up the professor courses management frame"""
        # Title
        ttk.Label(self.professor_courses_frame, text="Professor Course Management",
                 style='Heading.TLabel').pack(pady=(0, 20))

        # Create a split view with professors on the left and courses on the right
        main_pane = ttk.PanedWindow(self.professor_courses_frame, orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH, expand=True, pady=10)

        # Left frame for professor selection
        left_frame = ttk.Frame(main_pane, style='TFrame')
        main_pane.add(left_frame, weight=1)

        # Professor selection
        ttk.Label(left_frame, text="Select Professor:", style='SubHeading.TLabel').pack(anchor=tk.W, pady=(0, 10))

        # Create a frame for the professor list
        prof_list_frame = ttk.Frame(left_frame, style='TFrame')
        prof_list_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        # Create a treeview for the professors
        self.professors_tree = ttk.Treeview(prof_list_frame, columns=("id", "name", "branch"), show="headings")
        self.professors_tree.heading("id", text="ID")
        self.professors_tree.heading("name", text="Professor Name")
        self.professors_tree.heading("branch", text="Branch")

        self.professors_tree.column("id", width=50)
        self.professors_tree.column("name", width=200)
        self.professors_tree.column("branch", width=150)

        self.professors_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Add a scrollbar
        scrollbar = ttk.Scrollbar(prof_list_frame, orient=tk.VERTICAL, command=self.professors_tree.yview)
        self.professors_tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Bind selection event
        self.professors_tree.bind("<<TreeviewSelect>>", self.on_professor_selected)

        # Refresh button
        ttk.Button(left_frame, text="Refresh Professors",
                  command=self.refresh_professors_list).pack(fill=tk.X, pady=10)

        # Right frame for courses
        right_frame = ttk.Frame(main_pane, style='TFrame')
        main_pane.add(right_frame, weight=2)

        # Create a notebook for tabs
        self.prof_courses_notebook = ttk.Notebook(right_frame)
        self.prof_courses_notebook.pack(fill=tk.BOTH, expand=True)

        # Current courses tab
        current_courses_tab = ttk.Frame(self.prof_courses_notebook, style='TFrame')
        self.prof_courses_notebook.add(current_courses_tab, text="Current Courses")

        # Create a treeview for current courses
        self.current_courses_tree = ttk.Treeview(current_courses_tab, columns=("code", "name", "branch"), show="headings")
        self.current_courses_tree.heading("code", text="Course Code")
        self.current_courses_tree.heading("name", text="Course Name")
        self.current_courses_tree.heading("branch", text="Branch")

        self.current_courses_tree.column("code", width=100)
        self.current_courses_tree.column("name", width=250)
        self.current_courses_tree.column("branch", width=150)

        self.current_courses_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Add a scrollbar
        scrollbar = ttk.Scrollbar(current_courses_tab, orient=tk.VERTICAL, command=self.current_courses_tree.yview)
        self.current_courses_tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Remove course button
        ttk.Button(current_courses_tab, text="Remove Selected Course",
                  command=self.remove_course_from_selected_professor).pack(fill=tk.X, pady=10)

        # Available courses tab
        available_courses_tab = ttk.Frame(self.prof_courses_notebook, style='TFrame')
        self.prof_courses_notebook.add(available_courses_tab, text="Available Courses")

        # Create a treeview for available courses
        self.available_courses_tree = ttk.Treeview(available_courses_tab, columns=("code", "name"), show="headings")
        self.available_courses_tree.heading("code", text="Course Code")
        self.available_courses_tree.heading("name", text="Course Name")

        self.available_courses_tree.column("code", width=100)
        self.available_courses_tree.column("name", width=300)

        self.available_courses_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Add a scrollbar
        scrollbar = ttk.Scrollbar(available_courses_tab, orient=tk.VERTICAL, command=self.available_courses_tree.yview)
        self.available_courses_tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Add course button
        ttk.Button(available_courses_tab, text="Add Selected Course",
                  command=self.add_course_to_selected_professor).pack(fill=tk.X, pady=10)

    def setup_professors_frame(self):
        """Set up the professor schedules frame"""
        # Title
        ttk.Label(self.professors_frame, text="Professor Schedules",
                 style='Heading.TLabel').pack(pady=(0, 20))

        # Professor selection with improved layout
        select_frame = ttk.Frame(self.professors_frame, style='TFrame')
        select_frame.pack(fill=tk.X, pady=10)

        # Create a grid layout for better organization
        select_frame.columnconfigure(0, weight=0)  # Label
        select_frame.columnconfigure(1, weight=1)  # Combobox
        select_frame.columnconfigure(2, weight=0)  # Buttons

        # Row 1: Professor selection
        ttk.Label(select_frame, text="Select Professor:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.professor_var = tk.StringVar()
        self.professor_combo = ttk.Combobox(select_frame, textvariable=self.professor_var, width=30)
        self.professor_combo.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W+tk.E)

        # Button frame for row 1
        button_frame = ttk.Frame(select_frame, style='TFrame')
        button_frame.grid(row=0, column=2, padx=5, pady=5, sticky=tk.E)

        ttk.Button(button_frame, text="View Schedule",
                  command=self.view_professor_schedule).pack(side=tk.LEFT, padx=2)
        ttk.Button(button_frame, text="Refresh",
                  command=self.refresh_professors).pack(side=tk.LEFT, padx=2)

        # Row 2: Export button
        export_frame = ttk.Frame(self.professors_frame, style='TFrame')
        export_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(export_frame, text="Export All Schedules to Excel",
                  command=self.export_professor_schedule_to_excel).pack(anchor=tk.E, padx=5)

        # Schedule display with improved styling
        schedule_frame = ttk.LabelFrame(self.professors_frame, text="Professor Schedule", style='TFrame')
        schedule_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        # Use a monospaced font for better schedule display
        self.professor_text = scrolledtext.ScrolledText(
            schedule_frame,
            wrap=tk.WORD,
            font=('Courier New', 10)
        )
        self.professor_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Initial message with better formatting
        self.professor_text.insert(tk.END,
            "No professor schedules available. Please generate a timetable first using the Generate tab.\n\n"
            "Once a timetable is generated, you can view the schedule for a selected professor."
        )
    
    # Database functions
    def initialize_db(self):
        """Initialize the database with sample data"""
        try:
            initialize_database()
            self.console.insert(tk.END, "Database initialized successfully!\n")
            self.status_var.set("Database initialized")
            self.refresh_branches()
            self.refresh_courses()
            self.refresh_mapping_lists()
            self.refresh_courses_for_classes()
            self.update_stats()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to initialize database: {str(e)}")
            self.console.insert(tk.END, f"Error initializing database: {str(e)}\n")
            self.status_var.set("Database initialization failed")
    
    def update_stats(self):
        """Update the statistics on the home screen"""
        try:
            with _connect() as conn:
                cursor = conn.cursor()
                
                # Count branches
                cursor.execute("SELECT COUNT(*) FROM branches")
                branch_count = cursor.fetchone()[0]
                self.branch_count.set(str(branch_count))
                
                # Count courses
                cursor.execute("SELECT COUNT(*) FROM courses")
                course_count = cursor.fetchone()[0]
                self.course_count.set(str(course_count))
                
                # Count teachers
                cursor.execute("SELECT COUNT(*) FROM teachers")
                teacher_count = cursor.fetchone()[0]
                self.teacher_count.set(str(teacher_count))
        except Exception as e:
            print(f"Error updating stats: {str(e)}")
    
    def debug_database(self):
        """Debug the database contents"""
        # Redirect stdout to the database console
        old_stdout = sys.stdout
        sys.stdout = StdoutRedirector(self.db_console)
        
        try:
            # Clear the console first
            self.db_console.delete(1.0, tk.END)
            
            debug_database()
            self.status_var.set("Database debug completed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to debug database: {str(e)}")
            self.db_console.insert(tk.END, f"Error debugging database: {str(e)}\n")
            self.status_var.set("Database debug failed")
        finally:
            # Restore stdout
            sys.stdout = old_stdout
    
    def verify_database(self):
        """Verify the database integrity"""
        # Redirect stdout to the database console
        old_stdout = sys.stdout
        sys.stdout = StdoutRedirector(self.db_console)
        
        try:
            # Clear the console first
            self.db_console.delete(1.0, tk.END)
            
            result = verify_database_integrity()
            if result:
                self.status_var.set("Database verification passed")
            else:
                self.status_var.set("Database verification failed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to verify database: {str(e)}")
            self.db_console.insert(tk.END, f"Error verifying database: {str(e)}\n")
            self.status_var.set("Database verification failed")
        finally:
            # Restore stdout
            sys.stdout = old_stdout
    
    def refresh_branches(self):
        """Refresh the branch comboboxes"""
        try:
            with _connect() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT branch_name FROM branches ORDER BY branch_name")
                branches = [branch[0] for branch in cursor.fetchall()]
                
                # Update all branch comboboxes
                self.prof_branch_combo['values'] = branches
                self.course_branch_combo['values'] = branches
                self.mapping_branch_combo['values'] = branches
                
                if branches:
                    self.prof_branch_combo.current(0)
                    self.course_branch_combo.current(0)
                    self.mapping_branch_combo.current(0)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh branches: {str(e)}")
    
    def add_branch(self):
        """Add a new branch to the database"""
        branch_name = self.branch_name_var.get().strip()
        if not branch_name:
            messagebox.showerror("Error", "Branch name cannot be empty")
            return
        
        try:
            # Redirect stdout to the database console
            old_stdout = sys.stdout
            sys.stdout = StdoutRedirector(self.db_console)
            
            add_branch(branch_name)
            self.branch_name_var.set("")  # Clear the entry
            self.refresh_branches()
            self.update_stats()
            self.status_var.set(f"Branch '{branch_name}' added")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add branch: {str(e)}")
            self.db_console.insert(tk.END, f"Error adding branch: {str(e)}\n")
        finally:
            # Restore stdout
            sys.stdout = old_stdout
    
    def add_professor(self):
        """Add a new professor to the database"""
        prof_name = self.prof_name_var.get().strip()
        branch_name = self.prof_branch_var.get()
        
        if not prof_name:
            messagebox.showerror("Error", "Professor name cannot be empty")
            return
        
        if not branch_name:
            messagebox.showerror("Error", "Please select a branch")
            return
        
        try:
            # Redirect stdout to the database console
            old_stdout = sys.stdout
            sys.stdout = StdoutRedirector(self.db_console)
            
            add_professor(prof_name, branch_name)
            self.prof_name_var.set("")  # Clear the entry
            self.update_stats()
            self.status_var.set(f"Professor '{prof_name}' added")
            self.refresh_mapping_lists()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add professor: {str(e)}")
            self.db_console.insert(tk.END, f"Error adding professor: {str(e)}\n")
        finally:
            # Restore stdout
            sys.stdout = old_stdout
    
    # Course management functions
    def add_course(self):
        """Add a new course to the database"""
        course_code = self.course_code_var.get().strip().upper()
        course_name = self.course_name_var.get().strip()
        branch_name = self.course_branch_var.get()
        
        if not course_code or not course_name:
            messagebox.showerror("Error", "Course code and name cannot be empty")
            return
        
        if not branch_name:
            messagebox.showerror("Error", "Please select a branch")
            return
        
        try:
            # Redirect stdout to the course console
            old_stdout = sys.stdout
            sys.stdout = StdoutRedirector(self.course_console)
            
            # Add course with branch
            with _connect() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO courses (course_code, course_name, branch_name) VALUES (?, ?, ?)",
                    (course_code, course_name, branch_name)
                )
                
                # Initialize course_classes with default value (2 classes per week)
                cursor.execute(
                    "INSERT INTO course_classes (course_code, num_classes) VALUES (?, ?)",
                    (course_code, 2)
                )
                
                conn.commit()
                self.course_console.insert(tk.END, f"✅ Course '{course_code}' added successfully!\n")
            
            self.course_code_var.set("")  # Clear the entries
            self.course_name_var.set("")
            self.update_stats()
            self.status_var.set(f"Course '{course_code}' added")
            self.refresh_courses()
            self.refresh_mapping_lists()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add course: {str(e)}")
            self.course_console.insert(tk.END, f"Error adding course: {str(e)}\n")
        finally:
            # Restore stdout
            sys.stdout = old_stdout
    
    def refresh_courses(self):
        """Refresh the course comboboxes"""
        try:
            with _connect() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT course_code, course_name FROM courses ORDER BY course_code")
                courses = cursor.fetchall()
                
                # Format courses for display
                course_display = [f"{code}: {name}" for code, name in courses]
                
                # Update course comboboxes
                self.delete_course_combo['values'] = course_display
                self.mapping_course_combo['values'] = course_display
                
                if course_display:
                    self.delete_course_combo.current(0)
                    self.mapping_course_combo.current(0)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh courses: {str(e)}")
    
    def refresh_mapping_lists(self):
        """Refresh all lists for the mapping tab"""
        try:
            # Refresh branches
            self.refresh_branches()
            
            # Refresh courses
            self.refresh_courses()
            
            # Refresh teachers
            with _connect() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT teacher_name FROM teachers ORDER BY teacher_name")
                teachers = [teacher[0] for teacher in cursor.fetchall()]
                
                self.mapping_teacher_combo['values'] = teachers
                
                if teachers:
                    self.mapping_teacher_combo.current(0)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh mapping lists: {str(e)}")
    
    def add_mapping(self):
        """Add a branch-teacher-course mapping"""
        branch_name = self.mapping_branch_var.get()
        teacher_name = self.mapping_teacher_var.get()
        course_full = self.mapping_course_var.get()
        
        if not branch_name or not teacher_name or not course_full:
            messagebox.showerror("Error", "Please select branch, teacher, and course")
            return
        
        # Extract course code from the display string (e.g., "CS101: Introduction to Programming")
        course_code = course_full.split(":")[0].strip()
        
        try:
            # Redirect stdout to the course console
            old_stdout = sys.stdout
            sys.stdout = StdoutRedirector(self.course_console)
            
            add_branch_teacher_course_mapping(branch_name, teacher_name, course_code)
            self.status_var.set(f"Mapping added for {course_code}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add mapping: {str(e)}")
            self.course_console.insert(tk.END, f"Error adding mapping: {str(e)}\n")
        finally:
            # Restore stdout
            sys.stdout = old_stdout
    
    def delete_course(self):
        """Delete a course from the database"""
        course_full = self.delete_course_var.get()
        
        if not course_full:
            messagebox.showerror("Error", "Please select a course to delete")
            return
        
        # Extract course code from the display string
        course_code = course_full.split(":")[0].strip()
        
        # Confirm deletion
        confirm = messagebox.askyesno("Confirm Deletion", 
                                     f"Are you sure you want to delete course '{course_code}'?\n"
                                     "This will also delete all mappings for this course.")
        if not confirm:
            return
        
        try:
            # Redirect stdout to the course console
            old_stdout = sys.stdout
            sys.stdout = StdoutRedirector(self.course_console)
            
            delete_course(course_code)
            self.update_stats()
            self.status_var.set(f"Course '{course_code}' deleted")
            self.refresh_courses()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete course: {str(e)}")
            self.course_console.insert(tk.END, f"Error deleting course: {str(e)}\n")
        finally:
            # Restore stdout
            sys.stdout = old_stdout
    
    # Timetable generation functions
    def generate_timetable(self):
        """Generate a timetable using the genetic algorithm"""
        # Verify database integrity first
        try:
            if not verify_database_integrity():
                messagebox.showerror("Error", "Database integrity check failed. Please fix the issues before generating a timetable.")
                return
        except Exception as e:
            messagebox.showerror("Error", f"Failed to verify database: {str(e)}")
            return

        # Get branches
        try:
            with _connect() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT branch_name FROM branches")
                branches = [branch[0] for branch in cursor.fetchall()]

                if not branches:
                    messagebox.showerror("Error", "No branches found. Please add branches first.")
                    return
        except Exception as e:
            messagebox.showerror("Error", f"Failed to get branches: {str(e)}")
            return

        # Get generations
        try:
            generations = int(self.generations_var.get())
            if generations < 30:
                messagebox.showerror("Error", "Generations must be at least 30")
                return
            # Limit to 200 for better performance
            if generations > 200:
                if not messagebox.askyesno("Warning", "Using more than 200 generations may cause the application to become unresponsive. Continue anyway?"):
                    generations = 200
                    self.generations_var.set("200")
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number for generations")
            return

        # Show progress dialog
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Generating Timetable")
        progress_window.geometry("400x150")
        progress_window.transient(self.root)
        progress_window.grab_set()

        ttk.Label(progress_window, text="Generating timetable...", font=("Arial", 12)).pack(pady=10)
        progress = ttk.Progressbar(progress_window, mode="indeterminate")
        progress.pack(fill=tk.X, padx=20, pady=10)
        progress.start()

        status_var = tk.StringVar(value="Starting genetic algorithm...")
        ttk.Label(progress_window, textvariable=status_var).pack(pady=10)

        # Update the progress variable
        self.progress_var.set("Generating...")

        # Function to run in a separate thread
        def generate():
            try:
                # Update status in the progress window
                def update_status(msg):
                    self.root.after(0, lambda: status_var.set(msg))
                    self.root.after(0, lambda: self.progress_var.set(msg))

                # Update status to show progress
                update_status("Initializing genetic algorithm...")

                # Generate timetable using the genetic algorithm
                # Suppress verbose output by setting verbose=False
                result = genetic_algorithm(branches=branches, generations=generations, verbose=False)

                # Update status to show completion
                update_status("Timetable generation complete!")

                if result:
                    schedule, professor_schedules = result
                    # Store for later access
                    self.last_timetable = schedule
                    self.last_professor_schedules = professor_schedules

                    # Display timetable
                    timetable_output = io.StringIO()
                    print_timetable(schedule, file=timetable_output)
                    timetable_text = timetable_output.getvalue()

                    # Update UI in the main thread
                    self.root.after(0, lambda: self.update_timetable_display(timetable_text))
                    self.root.after(0, lambda: self.refresh_professors())
                    self.root.after(0, lambda: self.status_var.set("Timetable generated successfully"))
                    self.root.after(0, lambda: self.progress_var.set("Completed"))
                    self.root.after(0, lambda: progress_window.destroy())
                else:
                    self.root.after(0, lambda: messagebox.showerror("Error", "Failed to generate a valid timetable."))
                    self.root.after(0, lambda: self.progress_var.set("Failed"))
                    self.root.after(0, lambda: progress_window.destroy())
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Failed to generate timetable: {str(e)}"))
                self.root.after(0, lambda: self.progress_var.set("Error"))
                self.root.after(0, lambda: progress_window.destroy())

        # Start generation in a separate thread
        threading.Thread(target=generate, daemon=True).start()
    
    def update_timetable_display(self, timetable_text):
        """Update the timetable display with the generated timetable"""
        self.timetable_text.delete(1.0, tk.END)
        self.timetable_text.insert(tk.END, timetable_text)
        # Switch to the timetable tab
        self.show_timetable()
    
    # Professor schedule functions
    def refresh_professors(self):
        """Refresh the professor combobox"""
        try:
            # Get all professors who have schedules
            if self.last_professor_schedules:
                professors = list(self.last_professor_schedules.keys())
                professors.sort()

                self.professor_combo['values'] = professors

                if professors:
                    self.professor_combo.current(0)
            else:
                self.professor_combo['values'] = []
                self.professor_text.delete(1.0, tk.END)
                self.professor_text.insert(tk.END, "No professor schedules available. Please generate a timetable first.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh professors: {str(e)}")

    def refresh_professors_list(self):
        """Refresh the list of professors in the professor courses tab"""
        try:
            # Clear the treeview
            for item in self.professors_tree.get_children():
                self.professors_tree.delete(item)

            # Get all professors
            professors = get_all_professors()

            # Add professors to the treeview
            for prof in professors:
                self.professors_tree.insert("", "end", values=prof)

            # Update the status
            if professors:
                self.status_var.set(f"Found {len(professors)} professors")
            else:
                self.status_var.set("No professors found")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh professors list: {str(e)}")

    def on_professor_selected(self, event):
        """Handle professor selection in the treeview"""
        # Get the selected professor
        selection = self.professors_tree.selection()
        if not selection:
            return

        # Get the professor ID
        prof_id = self.professors_tree.item(selection[0], "values")[0]

        # Refresh the courses for this professor
        self.refresh_professor_courses(prof_id)

    def refresh_professor_courses(self, professor_id):
        """Refresh the courses for a professor"""
        try:
            # Clear the current courses treeview
            for item in self.current_courses_tree.get_children():
                self.current_courses_tree.delete(item)

            # Clear the available courses treeview
            for item in self.available_courses_tree.get_children():
                self.available_courses_tree.delete(item)

            # Get current courses
            current_courses = get_professor_courses(professor_id)

            # Add current courses to the treeview
            for course in current_courses:
                self.current_courses_tree.insert("", "end", values=course)

            # Get available courses
            available_courses = get_available_courses_for_professor(professor_id)

            # Add available courses to the treeview
            for course in available_courses:
                self.available_courses_tree.insert("", "end", values=course)

            # Update the notebook tab text to show counts
            self.prof_courses_notebook.tab(0, text=f"Current Courses ({len(current_courses)})")
            self.prof_courses_notebook.tab(1, text=f"Available Courses ({len(available_courses)})")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh professor courses: {str(e)}")

    def add_course_to_selected_professor(self):
        """Add a course to the selected professor"""
        # Get the selected professor
        prof_selection = self.professors_tree.selection()
        if not prof_selection:
            messagebox.showwarning("Warning", "Please select a professor first")
            return

        # Get the selected course
        course_selection = self.available_courses_tree.selection()
        if not course_selection:
            messagebox.showwarning("Warning", "Please select a course to add")
            return

        # Get the professor ID and course code
        prof_id = self.professors_tree.item(prof_selection[0], "values")[0]
        course_code = self.available_courses_tree.item(course_selection[0], "values")[0]

        # Add the course to the professor
        if add_course_to_professor(prof_id, course_code):
            messagebox.showinfo("Success", "Course added to professor successfully")
            # Refresh the courses
            self.refresh_professor_courses(prof_id)
        else:
            messagebox.showerror("Error", "Failed to add course to professor")

    def remove_course_from_selected_professor(self):
        """Remove a course from the selected professor"""
        # Get the selected professor
        prof_selection = self.professors_tree.selection()
        if not prof_selection:
            messagebox.showwarning("Warning", "Please select a professor first")
            return

        # Get the selected course
        course_selection = self.current_courses_tree.selection()
        if not course_selection:
            messagebox.showwarning("Warning", "Please select a course to remove")
            return

        # Get the professor ID and course code
        prof_id = self.professors_tree.item(prof_selection[0], "values")[0]
        course_code = self.current_courses_tree.item(course_selection[0], "values")[0]

        # Confirm removal
        if not messagebox.askyesno("Confirm", "Are you sure you want to remove this course from the professor?"):
            return

        # Remove the course from the professor
        if remove_course_from_professor(prof_id, course_code):
            messagebox.showinfo("Success", "Course removed from professor successfully")
            # Refresh the courses
            self.refresh_professor_courses(prof_id)
        else:
            messagebox.showerror("Error", "Failed to remove course from professor")
    
    # Course classes management functions
    def refresh_courses_for_classes(self):
        """Refresh the courses combobox for the classes tab"""
        try:
            with _connect() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT c.course_code, c.course_name, cc.num_classes
                    FROM courses c
                    JOIN course_classes cc ON c.course_code = cc.course_code
                    ORDER BY c.course_code
                """)
                courses = cursor.fetchall()

                # Format courses for display
                course_display = [f"{code}: {name} ({classes} classes/week)"
                                 for code, name, classes in courses]

                # Update course combobox
                self.update_course_combo['values'] = course_display

                if course_display:
                    self.update_course_combo.current(0)
                    # Set the current number of classes
                    if courses:
                        self.update_classes_var.set(str(courses[0][2]))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh courses: {str(e)}")

    def view_course_classes(self):
        """Display the course classes configuration"""
        # Redirect stdout to the classes console
        old_stdout = sys.stdout
        sys.stdout = StdoutRedirector(self.classes_console)

        try:
            # Clear the console first
            self.classes_console.delete(1.0, tk.END)

            # Display course classes
            course_classes = get_course_classes()

            if not course_classes:
                self.classes_console.insert(tk.END, "No courses found in the database.\n")
                return

            # Prepare data for display
            headers = ["Course Code", "Course Name", "Branch", "Classes per Week"]
            table_data = []

            for course_code, num_classes, course_name, branch_name in course_classes:
                table_data.append([course_code, course_name, branch_name, num_classes])

            # Print the table
            from tabulate import tabulate
            self.classes_console.insert(tk.END, "\n=== Course Classes Configuration ===\n")
            self.classes_console.insert(tk.END, tabulate(table_data, headers=headers, tablefmt="grid"))
            self.classes_console.insert(tk.END, f"\nTotal courses: {len(course_classes)}\n")

            self.status_var.set("Course classes displayed")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to view course classes: {str(e)}")
            self.classes_console.insert(tk.END, f"Error viewing course classes: {str(e)}\n")
        finally:
            # Restore stdout
            sys.stdout = old_stdout

    def update_course_classes(self):
        """Update the number of classes for a specific course"""
        course_full = self.update_course_var.get()

        if not course_full:
            messagebox.showerror("Error", "Please select a course")
            return

        # Extract course code from the display string
        course_code = course_full.split(":")[0].strip()

        try:
            num_classes = int(self.update_classes_var.get())
            if num_classes < 1:
                messagebox.showerror("Error", "Number of classes must be at least 1")
                return
            if num_classes > 5:
                confirm = messagebox.askyesno("Confirm",
                                             f"Are you sure you want to set {num_classes} classes per week? "
                                             "This might be difficult to schedule.")
                if not confirm:
                    return
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number")
            return

        try:
            # Redirect stdout to the classes console
            old_stdout = sys.stdout
            sys.stdout = StdoutRedirector(self.classes_console)

            if set_course_classes(course_code, num_classes):
                self.classes_console.insert(tk.END, f"✅ Successfully updated {course_code} to {num_classes} classes per week.\n")
                self.status_var.set(f"Updated {course_code} to {num_classes} classes/week")
                self.refresh_courses_for_classes()
            else:
                self.classes_console.insert(tk.END, f"❌ Failed to update {course_code}.\n")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update course classes: {str(e)}")
            self.classes_console.insert(tk.END, f"Error updating course classes: {str(e)}\n")
        finally:
            # Restore stdout
            sys.stdout = old_stdout

    def update_all_course_classes(self):
        """Update the number of classes for all courses"""
        try:
            num_classes = int(self.update_all_classes_var.get())
            if num_classes < 1:
                messagebox.showerror("Error", "Number of classes must be at least 1")
                return
            if num_classes > 5:
                confirm = messagebox.askyesno("Confirm",
                                             f"Are you sure you want to set {num_classes} classes per week for ALL courses? "
                                             "This might be difficult to schedule.")
                if not confirm:
                    return
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number")
            return

        # Confirm
        confirm = messagebox.askyesno("Confirm",
                                     f"This will set {num_classes} classes per week for ALL courses. Continue?")
        if not confirm:
            return

        try:
            # Redirect stdout to the classes console
            old_stdout = sys.stdout
            sys.stdout = StdoutRedirector(self.classes_console)

            # Update all courses
            course_classes = get_course_classes()
            success_count = 0

            for course_code, _, _, _ in course_classes:
                if set_course_classes(course_code, num_classes):
                    success_count += 1

            self.classes_console.insert(tk.END, f"✅ Successfully updated {success_count} out of {len(course_classes)} courses.\n")
            self.status_var.set(f"Updated all courses to {num_classes} classes/week")
            self.refresh_courses_for_classes()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update all course classes: {str(e)}")
            self.classes_console.insert(tk.END, f"Error updating all course classes: {str(e)}\n")
        finally:
            # Restore stdout
            sys.stdout = old_stdout

    def view_professor_schedule(self):
        """Display the schedule for the selected professor"""
        professor = self.professor_var.get()

        if not professor:
            messagebox.showerror("Error", "Please select a professor")
            return

        if not self.last_professor_schedules or professor not in self.last_professor_schedules:
            messagebox.showerror("Error", "No schedule available for this professor")
            return

        self.status_var.set(f"Schedule ready for {professor} - use Export to Excel to view")

    def export_timetable_to_excel(self):
        """Export the timetable to an Excel file"""
        if not self.last_timetable:
            messagebox.showwarning("Warning", "No timetable available to export")
            return

        # Ask for the file location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Timetable as Excel"
        )

        if not file_path:
            return  # User cancelled

        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Timetable"

            # Define styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            day_font = Font(bold=True, size=11)
            day_fill = PatternFill(start_color="E8F4F9", end_color="E8F4F9", fill_type="solid")

            time_font = Font(bold=True, size=10)
            time_fill = PatternFill(start_color="F2F9FC", end_color="F2F9FC", fill_type="solid")

            lunch_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Define days and time slots
            days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
            time_slots = ["8:00-9:00", "9:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-1:00",
                         "1:00-2:00", "2:00-3:00", "3:00-4:00", "4:00-5:00"]

            # Add title
            ws.merge_cells('A1:J1')
            ws['A1'] = "Timetable"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal="center")

            # Add generation date
            ws.merge_cells('A2:J2')
            ws['A2'] = f"Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ws['A2'].alignment = Alignment(horizontal="center")

            # Add headers
            ws['A4'] = "Time / Day"
            ws['A4'].font = header_font
            ws['A4'].fill = header_fill
            ws['A4'].alignment = header_alignment
            ws['A4'].border = border

            for i, day in enumerate(days):
                col = chr(66 + i)  # B, C, D, E, F
                ws[f'{col}4'] = day
                ws[f'{col}4'].font = header_font
                ws[f'{col}4'].fill = header_fill
                ws[f'{col}4'].alignment = header_alignment
                ws[f'{col}4'].border = border

            # Set column widths
            ws.column_dimensions['A'].width = 15
            for i in range(5):  # B to F
                col = chr(66 + i)
                ws.column_dimensions[col].width = 25

            # Add time slots and timetable data
            for i, time_slot in enumerate(time_slots):
                row = 5 + i

                # Time slot
                ws[f'A{row}'] = time_slot
                ws[f'A{row}'].font = time_font
                ws[f'A{row}'].fill = time_fill
                ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
                ws[f'A{row}'].border = border

                # Special handling for lunch slot
                if i == 4:  # Lunch slot
                    for j in range(5):  # B to F
                        col = chr(66 + j)
                        ws[f'{col}{row}'] = "LUNCH"
                        ws[f'{col}{row}'].fill = lunch_fill
                        ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center")
                        ws[f'{col}{row}'].border = border
                    continue

                # Add timetable data
                for j, day in enumerate(days):
                    col = chr(66 + j)
                    cell_value = ""

                    # Get the class data for this slot
                    class_data = self.last_timetable[j][i]

                    if class_data is not None:
                        if isinstance(class_data, list):
                            # Multiple classes in this slot
                            for cls in class_data:
                                if hasattr(cls, 'course_code') and cls.course_code != "LUNCH":
                                    cell_value += f"{cls.course_code}: {cls.course_name}\n"
                                    cell_value += f"Prof: {cls.teacher}\n"
                                    cell_value += f"Room: {cls.room}\n"
                                    cell_value += f"Branch: {cls.branch}\n\n"
                        elif hasattr(class_data, 'course_code') and class_data.course_code != "LUNCH":
                            # Single class in this slot
                            cell_value = f"{class_data.course_code}: {class_data.course_name}\n"
                            cell_value += f"Prof: {class_data.teacher}\n"
                            cell_value += f"Room: {class_data.room}\n"
                            cell_value += f"Branch: {class_data.branch}"

                    ws[f'{col}{row}'] = cell_value
                    ws[f'{col}{row}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    ws[f'{col}{row}'].border = border

            # Save the workbook
            wb.save(file_path)
            messagebox.showinfo("Success", f"Timetable exported successfully to {file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export timetable: {str(e)}")

    def export_professor_schedule_to_excel(self):
        """Export all professor schedules to an Excel file"""
        if not self.last_professor_schedules:
            messagebox.showwarning("Warning", "No professor schedules available to export")
            return

        # Ask for the file location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Professor Schedules as Excel"
        )

        if not file_path:
            return  # User cancelled

        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Professor Schedules"

            # Define styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            day_font = Font(bold=True, size=11)
            day_fill = PatternFill(start_color="E8F4F9", end_color="E8F4F9", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Set initial column widths
            ws.column_dimensions['A'].width = 15  # Day
            ws.column_dimensions['B'].width = 15  # Time
            ws.column_dimensions['C'].width = 30  # Course
            ws.column_dimensions['D'].width = 20  # Branch
            ws.column_dimensions['E'].width = 15  # Room

            # Write title
            ws.merge_cells('A1:E1')
            title_cell = ws['A1']
            title_cell.value = "Professor Schedules"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")

            current_row = 3  # Start from row 3

            # Write schedules for each professor
            for professor, schedule in self.last_professor_schedules.items():
                # Professor name as header
                ws.merge_cells(f'A{current_row}:E{current_row}')
                prof_cell = ws[f'A{current_row}']
                prof_cell.value = f"Schedule for {professor}"
                prof_cell.font = header_font
                prof_cell.fill = header_fill
                prof_cell.alignment = header_alignment
                
                current_row += 1

                # Headers for the schedule
                headers = ["Day", "Time", "Course", "Branch", "Room"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = day_font
                    cell.fill = day_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

                current_row += 1

                # Write schedule data
                days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
                time_slots = ["8:00-9:00", "9:00-10:00", "10:00-11:00", "11:00-12:00", 
                            "12:00-1:00", "1:00-2:00", "2:00-3:00", "3:00-4:00", "4:00-5:00"]

                for day in days:
                    for time in time_slots:
                        slot_key = f"{day} {time}"
                        if slot_key in schedule:
                            course_info = schedule[slot_key]
                            # Split the course info into components
                            parts = course_info.split(" - ")
                            if len(parts) >= 3:
                                course = parts[0]
                                branch = parts[1]
                                room = parts[2]
                            else:
                                course = course_info
                                branch = ""
                                room = ""

                            row_data = [day, time, course, branch, room]
                            for col, value in enumerate(row_data, 1):
                                cell = ws.cell(row=current_row, column=col)
                                cell.value = value
                                cell.border = border
                                cell.alignment = Alignment(horizontal="center")

                            current_row += 1

                current_row += 2  # Add space between professors

            # Save the workbook
            wb.save(file_path)
            messagebox.showinfo("Success", "Professor schedules exported successfully!")
            self.status_var.set("Professor schedules exported to Excel")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export professor schedules: {str(e)}")
            self.status_var.set("Error exporting professor schedules")
    
def main():
    root = tk.Tk()
    app = ModernTimetableApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()